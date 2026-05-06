import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from pathlib import Path
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table,
    TableStyle, Image
)

DIMS = ["helpfulness", "accuracy", "coherence", "tone", "overall"]

# ── Helpers ───────────────────────────────────────────────────────────────────

def score_color(score: float):
    if score >= 4:
        return colors.HexColor("#1D9E75")
    elif score >= 3:
        return colors.HexColor("#BA7517")
    else:
        return colors.HexColor("#E24B4A")

def generate_charts(df: pd.DataFrame, output_dir: str) -> dict:
    paths = {}

    # Bar chart — average scores per dimension
    fig, ax = plt.subplots(figsize=(7, 3.5))
    avgs = [df[d].mean() for d in DIMS]
    bar_colors = ["#1D9E75" if v >= 4 else "#BA7517" if v >= 3 else "#E24B4A" for v in avgs]
    bars = ax.bar([d.capitalize() for d in DIMS], avgs, color=bar_colors, width=0.5)
    ax.set_ylim(0, 5)
    ax.set_ylabel("Average Score")
    ax.set_title("Average Scores by Dimension")
    ax.axhline(y=4, color="gray", linestyle="--", linewidth=0.8, label="Pass threshold")
    for bar, val in zip(bars, avgs):
        ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.1,
                f"{val:.2f}", ha="center", va="bottom", fontsize=9)
    ax.legend()
    plt.tight_layout()
    bar_path = f"{output_dir}/chart_bar.png"
    plt.savefig(bar_path, dpi=150)
    plt.close()
    paths["bar"] = bar_path

    # Score distribution histogram
    fig, ax = plt.subplots(figsize=(7, 3))
    ax.hist(df["overall"], bins=[0.5,1.5,2.5,3.5,4.5,5.5],
            color="#378ADD", edgecolor="white", rwidth=0.8)
    ax.set_xlabel("Overall Score")
    ax.set_ylabel("Number of Responses")
    ax.set_title("Overall Score Distribution")
    ax.set_xticks([1, 2, 3, 4, 5])
    plt.tight_layout()
    hist_path = f"{output_dir}/chart_hist.png"
    plt.savefig(hist_path, dpi=150)
    plt.close()
    paths["hist"] = hist_path

    return paths

# ── Excel Export ──────────────────────────────────────────────────────────────

def _load(results_path: str) -> pd.DataFrame:
    df = pd.read_csv(results_path)
    for d in DIMS:
        if d in df.columns:
            df[d] = pd.to_numeric(df[d], errors="coerce").fillna(0).astype(int)
    return df


def export_excel(results_path: str, output_dir: str = "results") -> str:
    df = _load(results_path)
    output_path = f"{output_dir}/eval_report.xlsx"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # Sheet 1 — All results
        df.to_excel(writer, sheet_name="All Results", index=False)
        ws = writer.sheets["All Results"]

        # Color code score cells
        from openpyxl.styles import PatternFill, Font, Alignment
        green = PatternFill("solid", fgColor="E1F5EE")
        amber = PatternFill("solid", fgColor="FAEEDA")
        red   = PatternFill("solid", fgColor="FCEBEB")

        dim_cols = {dim: df.columns.get_loc(dim) + 1 for dim in DIMS}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for dim, col_idx in dim_cols.items():
                cell = row[col_idx - 1]
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 4:
                        cell.fill = green
                    elif cell.value >= 3:
                        cell.fill = amber
                    else:
                        cell.fill = red

        # Bold header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Sheet 2 — Summary
        summary_data = {
            "Dimension": [d.capitalize() for d in DIMS],
            "Mean Score": [round(df[d].mean(), 2) for d in DIMS],
            "Min": [int(df[d].min()) for d in DIMS],
            "Max": [int(df[d].max()) for d in DIMS],
            "Pass Rate (>=4)": [f"{(df[d] >= 4).mean() * 100:.0f}%" for d in DIMS]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        ws2 = writer.sheets["Summary"]
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Sheet 3 — Worst cases
        worst = df.nsmallest(5, "overall")[["id", "user_message", "ai_response", "overall", "reasoning"]]
        worst.to_excel(writer, sheet_name="Worst Cases", index=False)

        ws3 = writer.sheets["Worst Cases"]
        for cell in ws3[1]:
            cell.font = Font(bold=True)

    print(f"Excel report saved to {output_path}")
    return output_path

# ── PDF Export ────────────────────────────────────────────────────────────────

def export_pdf(results_path: str, output_dir: str = "results") -> str:
    df = _load(results_path)
    output_path = f"{output_dir}/eval_report.pdf"
    charts = generate_charts(df, output_dir)

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            rightMargin=0.75*inch, leftMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=20, spaceAfter=6)
    story.append(Paragraph("EvalAI - Evaluation Report", title_style))
    story.append(Paragraph(f"Total responses evaluated: <b>{len(df)}</b>", styles["Normal"]))
    story.append(Spacer(1, 0.2*inch))

    # Summary table
    story.append(Paragraph("Summary by Dimension", styles["Heading2"]))
    story.append(Spacer(1, 0.1*inch))

    table_data = [["Dimension", "Mean Score", "Min", "Max", "Pass Rate"]]
    for dim in DIMS:
        col = df[dim]
        table_data.append([
            dim.capitalize(),
            f"{col.mean():.2f}",
            str(int(col.min())),
            str(int(col.max())),
            f"{(col >= 4).mean() * 100:.0f}%"
        ])

    t = Table(table_data, colWidths=[1.5*inch, 1.2*inch, 0.8*inch, 0.8*inch, 1.2*inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C2C2A")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F1EFE8"), colors.white]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#B4B2A9")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3*inch))

    # Charts
    story.append(Paragraph("Score Analysis", styles["Heading2"]))
    story.append(Spacer(1, 0.1*inch))
    story.append(Image(charts["bar"], width=6*inch, height=3*inch))
    story.append(Spacer(1, 0.2*inch))
    story.append(Image(charts["hist"], width=6*inch, height=2.5*inch))
    story.append(Spacer(1, 0.3*inch))

    # Worst cases
    story.append(Paragraph("Worst Performing Cases", styles["Heading2"]))
    story.append(Spacer(1, 0.1*inch))

    worst = df.nsmallest(3, "overall")
    for _, row in worst.iterrows():
        story.append(Paragraph(
            f"<b>[{row['id']}]</b> Overall: {row['overall']}/5",
            styles["Heading3"]
        ))
        story.append(Paragraph(f"<b>Question:</b> {str(row['user_message'])}", styles["Normal"]))
        story.append(Paragraph(f"<b>AI Response:</b> {str(row['ai_response'])[:200]}...", styles["Normal"]))
        story.append(Paragraph(f"<b>Reasoning:</b> {str(row['reasoning'])}", styles["Normal"]))
        story.append(Spacer(1, 0.15*inch))

    doc.build(story)
    print(f"PDF report saved to {output_path}")
    return output_path